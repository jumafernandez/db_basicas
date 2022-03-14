# -*- coding: utf-8 -*-

from credenciales import USER, PASSWORD, HOST, PORT, DB

def get_data_pg(query, USER_DB, PASSWORD_DB, HOST_DB, PORT_DB, DB_NAME):
    """
    Parameters
    ----------
    query : TEXT
        Se pasa el query a correr en la db
    USER, PASSWORD, HOST, PORT, DB: TEXT
        Datos de conexión a la DB
    Returns
    -------
    data : list(tuples)
        Devuelve los datos que devuelve la db
    """

    import psycopg2
    import pandas as pd

    try:
        connection = psycopg2.connect(user=USER_DB,
                                      password=PASSWORD_DB,
                                      host=HOST_DB,
                                      port=PORT_DB,
                                      database=DB_NAME)

        data = pd.read_sql_query(query, connection)
        connection = None
        
    except (Exception, psycopg2.Error) as error:
        print("No es posible devolver los datos de la DB.", error)
    
    finally:
        # Se cierra la conexión a la DB
        if connection:
            connection.close()

    return data


def generacion_reportes(division_parametro, anio_parametro, dir_reportes, db_update):
    
    import pandas as pd
    
    # Recorro la lista de divisiones que me pasaron para generar los reportes
    for division in division_parametro:
        
        # Genero un path para el archivo
        path_archivo = f"{dir_reportes}/División-{division}-al-{db_update}.xlsx"
        
        # Defino el query con la info que quiero
        query_dim_docente = f"""SELECT * 
                                FROM reporte_dimension_docente_anual 
                                WHERE division='{division}' 
                                AND anio={anio_parametro}
                                ORDER BY 1;"""
        
        # Recupero la información de la base en función del query
        df_docente = get_data_pg(query_dim_docente, USER, PASSWORD, HOST, PORT, DB)
    
        # Defino el query con la info que quiero
        query_dim_servicios = f"""SELECT * 
                                FROM resumen_equipos_por_actividad 
                                WHERE division='{division}' 
                                AND anio_cursada={anio_parametro}
                                ORDER BY 1, 2, 3;"""
            
        # Recupero la información de la base en función del query
        df_servicios = get_data_pg(query_dim_servicios, USER, PASSWORD, HOST, PORT, DB)

        # Defino el query con la info que quiero
        query_dim_cargos = f"""SELECT *
                               FROM resumen_cargos_por_division 
                               WHERE division='{division}'
                               ORDER BY 1;"""
            
        # Recupero la información de la base en función del query
        df_cargos = get_data_pg(query_dim_cargos, USER, PASSWORD, HOST, PORT, DB)

        # Defino el query con la info que quiero
        query_dim_equipos = f"""SELECT *
                               FROM resumen_equipos_por_oferta_academica 
                               WHERE division='{division}' AND anio_cursada={anio_parametro}
                               ORDER BY 1;"""
            
        # Recupero la información de la base en función del query
        df_equipos = get_data_pg(query_dim_equipos, USER, PASSWORD, HOST, PORT, DB)
    
        # Genero el ExcelWriter
        writer = pd.ExcelWriter(path_archivo, engine = 'xlsxwriter')

        # Escribo al excel de la división
        df_cargos.to_excel(writer, sheet_name = f'Cargos al {db_update}', index=False)
        df_equipos.to_excel(writer, sheet_name = f'Equipos docentes - Año {anio_parametro}', index=False)
        df_docente.to_excel(writer, sheet_name = 'indicadores_docentes', index=False)
        df_servicios.to_excel(writer, sheet_name = 'indicadores_servicios', index=False)

        # Cierro el Excel writer
        writer.save()
               
def capa_presentacion_general(division_parametro, dir_reportes, db_update):
    """
    Función que trabaja sobre el formato de los reportes.

    Parameters
    ----------
    division_parametro : list
        Contiene las divisiones para las cuales se generan los reportes.
    dir_reportes : text
        Contiene el PATH donde están los reportes
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Recorro la lista de divisiones que me pasaron para generar los reportes
    for division in division_parametro:
        
        # Genero un path para el archivo
        path_archivo = f"{dir_reportes}/División-{division}-al-{db_update}.xlsx"

        libro = load_workbook(filename = path_archivo)
        
        for sheet_name in libro.sheetnames:

            sheet = libro[sheet_name]

            # Tomo la última fila-columna editadas
            row_count = sheet.max_row
            column_count = sheet.max_column
            print(f'División: {division}, Sheet: {sheet_name}.')
    
           # Defino los formatos de bordes
            thin = Side(border_style="thin", color="000000") # De cada lado
            border = Border(left=thin, right=thin, top=thin, bottom=thin) # de toda la celda
     
            i=0
            for fila in sheet.iter_rows(min_row=1, min_col=1, max_row=row_count, max_col=column_count):
                i = i + 1
                for celda in fila:
                    if i==1: # Si es la primera fila tiene formato distinto
                        celda.fill = PatternFill("solid", fgColor="DDEBF7")
                        celda.alignment = Alignment(horizontal="center", vertical="center")
                    else: # En caso que no sea la primera fila
                        # El color de fondo es blanco
                        celda.fill = PatternFill("solid", fgColor="FFFFFF")
                        # Está centrado si es un número
                        if (type(celda.value) is int) or (type(celda.value) is float): 
                            celda.alignment = Alignment(horizontal="center", vertical="center")
                            
                    # Todas las celdas tienen borde
                    celda.border = border
            
            # Acomodo el ancho óptimo de las columnas
            for column_cells in sheet.columns:
                new_column_length = max(len(str(cell.value)) for cell in column_cells)
                new_column_letter = (get_column_letter(column_cells[0].column))
                if new_column_length > 15:
                    sheet.column_dimensions[new_column_letter].width = new_column_length*1.02
                if new_column_length > 0:
                    sheet.column_dimensions[new_column_letter].width = new_column_length*1.25

                    
        # Guardo los cambios            
        libro.save(path_archivo)

def formato_condicional(division_parametro, dir_reportes, db_update):
    """
    Se incorpora formato condicional para aquellas columnas con criterios predefinidos

    Parameters
    ----------
    division_parametro : list
        Contiene las divisiones para las cuales se generan los reportes.
    dir_reportes : text
        Contiene el PATH donde están los reportes
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
#    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
    
    # Recorro la lista de divisiones que me pasaron para generar los reportes
    for division in division_parametro:
        
        # Genero un path para el archivo
        path_archivo = f"{dir_reportes}/División-{division}-al-{db_update}.xlsx"

        libro = load_workbook(filename = path_archivo)
        
        sheet_docentes = libro['indicadores_docentes']

        print(f'División: {division}, Sheet: indicadores_docentes.')
    
        # Creo los estilos para las reglas
        redFill = PatternFill(fgColor='FF8585', fill_type='solid')
        yellowFill = PatternFill(fgColor='FAFA8A', fill_type='solid')
        greenFill = PatternFill(fgColor='97FE86', fill_type='solid')

        # Recorro la columna J donde está el índice de cobertura docente
        i=1
        for celda in sheet_docentes['J']:
            
            # Si es la primera celda no intervengo
            if(i!=1):
                # En caso que haya un valor
                if celda.value is not None:
                    # Evalúo el color de la celda en función del valor
                    if celda.value >= 0.8:
                        celda.fill = yellowFill
                    elif celda.value < 0.5:
                        celda.fill = redFill
                    elif (celda.value >= 0.5 and celda.value < 0.8):
                        celda.fill = greenFill
                    # Defino como porcentaje y dos ceros al formato
                    celda.number_format = FORMAT_PERCENTAGE_00

            # Incremento 1 para saber en que fila estoy
            i=i+1
            
                
        # Guardo los cambios            
        libro.save(path_archivo)

if __name__ == '__main__':
    # Path del archivo a cargar
    DIVISIONES = ['Computación', 'Biología', 'Estadística', 'Física', 'Matemática', 'Química']
    ANIO = 2021
    DIRECTORIO_REPORTES = 'C:/Users/Juan/Documents/GitHub/db_basicas/app/reportes'
    ACTUALIZACION_DB = '2022-03-12'
    
    generacion_reportes(DIVISIONES, ANIO, DIRECTORIO_REPORTES, ACTUALIZACION_DB)
    capa_presentacion_general(DIVISIONES, DIRECTORIO_REPORTES, ACTUALIZACION_DB)
    formato_condicional(DIVISIONES, DIRECTORIO_REPORTES, ACTUALIZACION_DB)
    